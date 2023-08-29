# =-= coding:utf-8 =-=
import os
import json
from function import select
import openpyxl

res=select(["根据.event更新index.json和index.xlsx","根据index.xlsx更新.event和index.json"])
if res==0:
    events = {}
    path = "events"
    #excel初始化
    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.append(["id","文件","标题","条件","可跳过","描述","事件类型","触发类型"])

    for file in os.listdir(path):
        file_path = os.path.join(path, file)
        if os.path.isfile(file_path):
            if not ".event" in file_path:
                continue
            f = open(file_path,"r",encoding="utf-8")
            res = f.readlines()
            f.close()
            events[int(res[0][:-1])]=[int(res[0][:-1]),file_path,res[1][:-1],res[2][:-1],bool(eval(res[3][:-1])),res[4][:-1],res[5][:-1],res[6][:-1]]
            ws.append([int(res[0][:-1]),file_path,res[1][:-1],res[2][:-1],res[3][:-1],res[4][:-1]],res[5][:-1],res[6][:-1])
        else:
            for file1 in os.listdir(os.path.join(file_path)):
                file_path1 = os.path.join(file_path,file1)
                if os.path.isfile(file_path1):
                    if not ".event" in file_path1:
                        continue
                    f = open(file_path1,"r",encoding="utf-8")
                    res = f.readlines()
                    f.close()
                    events[int(res[0][:-1])]=[int(res[0][:-1]),file_path1,res[1][:-1],res[2][:-1],bool(eval(res[3][:-1])),res[4][:-1],res[5][:-1],res[6][:-1]]
                    ws.append([int(res[0][:-1]),file_path1,res[1][:-1],res[2][:-1],res[3][:-1],res[4][:-1],res[5][:-1],res[6][:-1]])
    f = open("events/index.json","w",encoding="utf-8")
    datar = json.dumps(events, sort_keys=True, indent=4, separators=(',', ': '))
    f.write(datar)
    f.close()
    #excel设置列宽
    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        max_width = 0
        for cell in ws[col_letter]:
            if len(str(cell.value)) > max_width:
                max_width = len(str(cell.value))
        if col_letter == 'A':
            max_width = max(max_width, 4)
        elif col_letter == 'E':
            max_width = max(max_width, 8)
        ws.column_dimensions[col_letter].width = max_width
    wb.save("events/index.xlsx")
elif res==1:
    wb = openpyxl.load_workbook("events/index.xlsx")
    ws=wb["Sheet"]
    events={}
      # 因为按行，所以返回A1, B1, C1这样的顺序
    for row in ws.rows:
        if row[0].value=='id':
            continue
        events[int(row[0].value)]=[int(row[0].value),row[1].value,row[2].value,row[3].value,bool(eval(row[4].value)),str(row[5].value),row[6][:-1],row[7][:-1]]
        f = open(row[1].value,"r+",encoding="utf-8")
        l = f.readlines()
        l[0] = str(row[0].value)+"\n"
        l[1] = row[2].value+"\n"
        l[2] = row[3].value+"\n"
        l[3] = row[4].value+"\n"
        l[4] = row[5].value+"\n"
        l[5] = row[6].value+"\n"
        l[6] = row[7].value+"\n"
        f.seek(0)
        f.writelines(l)
        f.close()
        
    f = open("events/index.json","w",encoding="utf-8")
    datar = json.dumps(events, sort_keys=True, indent=4, separators=(',', ': '))
    f.write(datar)
    f.close()
